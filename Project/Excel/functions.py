from collections import namedtuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from typing import NamedTuple
import datetime
import calendar
from datetime import timedelta
from datetime import date
from tkinter import messagebox

class Times(NamedTuple):
    startDate: datetime.date
    endDate: datetime.date

class Salary(NamedTuple):
    salary: float
    date: Times
    dailyHours: float
    contractDays: int
    daysAtSalary: int
    dailyRate: float
    hourlyRate: float

class Person(NamedTuple):
    eeNum: str
    name: str
    posit: str
    processLevel: str
    effDate: datetime.date
    reason: str
    union: str
    actionTaken: str
    sick: Times
    unearned: Times
    unpaid: Times
    currSalary: Salary
    salaries: list


class ExcelDoc(object):

    def __init__(self, workbook, destination, **kwags):
        self.wb = load_workbook(filename=workbook, read_only=True)
        self.ws = self.wb["Input"]
        self.dest = destination
        self.author = "John Doe" # change to read from input sheet
        self.firstPay = datetime.date.min
        self.lastPay = datetime.date.min
        self.today = datetime.date.today()
        self.grayFill = PatternFill(start_color='808080', end_color='808080', fill_type='solid');
        self.bd = Side(style='thick', color="000000");
        self.headerFont = Font(name='Arial', size=14);
        self.miniBlueFont = Font(name='Arial', size=8, color='0070C0');
        self.normalBlueFont = Font(name='Arial', size=10, color='0070C0', bold=True);
        self.whiteFont = Font(name='Arial', size=10, color='FFFFFF');
        self.bold = Font(name='Arial', size=10, bold=True);
        self.centerHorz = Alignment(horizontal='center');

    def numOfBusDaysInMonth(self, start, end, month):
        numOfDays = 0
        months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"];
        monthNum = months.index(month) + 1
        beginMonths = [7, 8, 9, 10, 11, 12];
        beginYear = start.year if start.month in beginMonths else start.year - 1;
        endYear = beginYear + 1;
        print("BeginYear: " + beginYear, "endYear: " + beginYear)
        year = beginYear if monthNum in beginMonths else endYear;
        firstDay = datetime.date(year, monthNum, 1 );
        lastDay = datetime.date(year, months.index(month) + 1, calendar.monthrange(year, monthNum)[1] );

        if not (lastDay < start or firstDay > end):
            if firstDay >= start:
                if lastDay <= end:
                    numOfDays = self.numOfBusDays(firstDay, lastDay);
                elif lastDay > end:
                    numOfDays = self.numOfBusDays(firstDay, end);
            elif firstDay < start:
                if lastDay <= end:
                    numOfDays = self.numOfBusDays(start, lastDay);
                elif lastDay > end:
                    numOfDays = self.numOfBusDays(start, end);
        return numOfDays;

    def numOfBusDays(self, start, end):
        daydiff = end.weekday() - start.weekday()
        days = ((end - start).days - daydiff) / 7 * 5 + min(daydiff, 5) - (max(end.weekday() - 4, 0) % 5)
        return days

    def positionCheck(self, position):
        position = position.lower()
        dailyHours = 0.0;
        contractDays = 0;
        union = "";
        if "inclusion" in position or "helper" in position or "ih" in position:
            dailyHours = 7.5;
            contractDays = 184;
            union = "HCESC"
            position = "Inclusion Helper"

        elif "new teacher" in position:
            dailyHours = 7.5;
            contractDays = 192;
            union = "HCEA"
            position = "Teacher"

        elif "custodian" in position:
            dailyHours = 8;
            contractDays = 260;
            union = "AFSCME"
            position = "Custodian"

        elif "10m" in position:
            dailyHours = 7.5;
            contractDays = 190;
            union = "HCEA"
            position = position.replace('10m', '');
            position = position.strip();
            position = position.title();

        elif "12m" in position:
            dailyHours = 7.5;
            contractDays = 260;
            union = "HCESC";
            position = position.replace('12m', '');
            position = position.strip();
            position = position.title();

        else:
            position = position.strip();
            position = position.title();

        return [dailyHours, contractDays, union, position];


    def produceForm(self, info):
        print(info);
        sick = Times(datetime.date.min, datetime.date.min);
        unearned = Times(datetime.date.min, datetime.date.min);
        unpaid = Times(datetime.date.min, datetime.date.min);
        if info[8] != "None" and info[9] != "None":
            sick = Times(datetime.datetime.strptime(info[8], "%Y-%m-%d %H:%M:%S").date(), datetime.datetime.strptime(info[9], "%Y-%m-%d %H:%M:%S").date());
        if info[10] != "None" and info[11] != "None":
            unearned = Times(datetime.datetime.strptime(info[10], "%Y-%m-%d %H:%M:%S").date(), datetime.datetime.strptime(info[11], "%Y-%m-%d %H:%M:%S").date());
        if info[12] != "None" and info[13] != "None":
            unpaid = Times(datetime.datetime.strptime(info[12], "%Y-%m-%d %H:%M:%S").date(), datetime.datetime.strptime(info[13], "%Y-%m-%d %H:%M:%S").date());
        posInfo = self.positionCheck(info[2]);
        d0 = datetime.datetime.strptime(info[14], "%Y-%m-%d %H:%M:%S").date();
        d1 = datetime.datetime.strptime(info[15], "%Y-%m-%d %H:%M:%S").date();
        contractDays = posInfo[1];
        dailyHours = posInfo[0];
        if info[17] != "None":
            contractDays = float(info[17]);
        if info[18] != "None":
            dailyHours = float(info[18]);
        currSalary = Salary(info[16], Times(d0, d1), posInfo[0], posInfo[1], self.numOfBusDays(d0,d1), float(info[16])/float(posInfo[1]), (float(info[16])/float(posInfo[1]))/float(posInfo[0]));
        union = info[6];
        salaries = [];
        if union == "None":
            union = posInfo[2];
        if len(info) > 19:
            for i in range(19, len(info), 5):
                d0 = datetime.datetime.strptime(info[i], "%Y-%m-%d %H:%M:%S").date();
                d1 = datetime.datetime.strptime(info[i+1], "%Y-%m-%d %H:%M:%S").date();
                if info[i+3] != "None":
                    contractDays = float(info[i+3]);
                if info[i+4] != "None":
                    dailyHours = float(info[i+4]);
                salaries.append(Salary(info[i+2], Times(d0, d1), dailyHours, contractDays, self.numOfBusDays(d0,d1), float(info[i+2]) / float(contractDays), (float(info[i+2]) / float(contractDays)) / float(dailyHours)));

        form = Person(info[0], info[1], posInfo[3], info[3], datetime.datetime.strptime(info[4], "%Y-%m-%d %H:%M:%S").date(), info[5], union, info[7], sick, unearned, unpaid, currSalary, salaries);
        print(form);
        self.setUpForm(form)


    def readFromInput(self):
        self.author = self.ws.cell(row=2, column=1).value;
        first = str(self.ws.cell(row=2, column=2).value).strip();
        last = str(self.ws.cell(row=2, column=3).value).strip();
        self.firstPay = datetime.datetime.strptime(first, "%Y-%m-%d %H:%M:%S").date();
        self.lastPay = datetime.datetime.strptime(last, "%Y-%m-%d %H:%M:%S").date();
        rowCount = self.ws.max_row;
        columnCount = self.ws.max_column;
        for row in range(5, rowCount + 1):
            fullRow = [];
            for col in range(1, columnCount + 1):
                cell = str(self.ws.cell(row=row, column=col).value).strip();
                if col > 19 and cell == "None" and (col-19)%6 == 0:
                    break;
                fullRow.append(cell);
            self.produceForm(fullRow);

    def setUpForm(self, person):
        wb = Workbook()
        ws = wb.active
        ws.title = "Form"
        for i in range(1, 14):
            col = ws.column_dimensions[get_column_letter(i)];
            col.width = 15.83;
        col = ws.column_dimensions[get_column_letter(6)];
        col.width = 1;
        col = ws.column_dimensions[get_column_letter(8)];
        col.width = 1.17;
        col = ws.column_dimensions[get_column_letter(10)];
        col.width = 2.67;
        col = ws.column_dimensions[get_column_letter(12)];
        col.width = 2.5;

        # row 1
        ws.cell(row=1, column=1).value = "HCPS Payroll Department: Contract Days vs Pay Cycle Days and Unearned Leave Calculation";
        ws.cell(row=1, column=1).font = self.headerFont;

        # row 2
        ws.cell(row=2, column=1).value = "Prepared By:";
        ws.cell(row=2, column=1).font = self.headerFont;
        ws.cell(row=2, column=2).font = self.headerFont;
        ws.cell(row=2, column=3).value = self.author;
        ws.cell(row=2, column=3).font = self.headerFont;
        ws.cell(row=2, column=4).font = self.headerFont;
        ws.cell(row=2, column=5).font = self.headerFont;
        ws.cell(row=2, column=6).font = self.headerFont;
        ws.cell(row=2, column=7).value = "Date:";
        ws.cell(row=2, column=7).font = self.headerFont;
        ws.cell(row=2, column=8).font = self.headerFont;
        ws.cell(row=2, column=9).value = self.today.strftime("%m/%d/%Y");
        ws.cell(row=2, column=9).font = self.headerFont;

        # row 3
        ws.cell(row=3, column=1).value = "Note: Cells in blue are produced";
        ws.cell(row=3, column=1).font = self.miniBlueFont;
        ws.cell(row=3, column=1).alignment = self.centerHorz;
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=13);

        # row 4
        ws.cell(row=4, column=1).value = "Information";
        ws.cell(row=4, column=1).fill = self.grayFill;
        ws.cell(row=4, column=1).font = self.whiteFont;
        ws.cell(row=4, column=1).alignment = self.centerHorz;
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=13);

        # row 5
        ws.cell(row=5, column=1).value = "EE#:";
        ws.cell(row=5, column=1).font = self.bold;
        ws.cell(row=5, column=2).value = int(person.eeNum);
        ws.cell(row=5, column=3).value = "Eff Date:";
        ws.cell(row=5, column=3).font = self.bold;
        ws.cell(row=5, column=4).value = person.effDate.strftime("%m/%d/%Y");
        ws.cell(row=5, column=5).value = "Sick Bank Grant:";
        ws.cell(row=5, column=5).font = self.bold;
        ws.cell(row=5, column=7).value = "Start Date:";
        ws.cell(row=5, column=7).font = self.bold;
        ws.cell(row=5, column=11).value = "End Date:";
        ws.cell(row=5, column=11).font = self.bold;
        if person.sick.startDate != datetime.date.min:
            ws.cell(row=5, column=8).value = person.sick.startDate.strftime("%m/%d/%Y");
            ws.cell(row=5, column=13).value = person.sick.endDate.strftime("%m/%d/%Y");

        # row 6
        ws.cell(row=6, column=1).value = "Name:";
        ws.cell(row=6, column=1).font = self.bold;
        ws.cell(row=6, column=2).value = person.name;
        ws.cell(row=6, column=3).value = "Reason:";
        ws.cell(row=6, column=3).font = self.bold;
        ws.cell(row=6, column=4).value = person.reason;
        ws.cell(row=6, column=5).value = "Unearned Leave:";
        ws.cell(row=6, column=5).font = self.bold;
        ws.cell(row=6, column=7).value = "Start Date:";
        ws.cell(row=6, column=7).font = self.bold;
        ws.cell(row=6, column=11).value = "End Date:";
        ws.cell(row=6, column=11).font = self.bold;
        if person.unearned.startDate != datetime.date.min:
            ws.cell(row=6, column=8).value = person.unearned.startDate.strftime("%m/%d/%Y");
            ws.cell(row=6, column=13).value = person.unearned.endDate.strftime("%m/%d/%Y");

        # row 7
        ws.cell(row=7, column=1).value = "Position:";
        ws.cell(row=7, column=1).font = self.bold;
        ws.cell(row=7, column=2).value = person.posit;
        ws.cell(row=7, column=3).value = "Union:";
        ws.cell(row=7, column=3).font = self.bold;
        ws.cell(row=7, column=4).value = person.union;
        ws.cell(row=7, column=5).value = "Unpaid:";
        ws.cell(row=7, column=5).font = self.bold;
        ws.cell(row=7, column=7).value = "Start Date:";
        ws.cell(row=7, column=7).font = self.bold;
        ws.cell(row=7, column=11).value = "End Date:";
        ws.cell(row=7, column=11).font = self.bold;
        if person.sick.startDate != datetime.date.min:
            ws.cell(row=7, column=8).value = person.unpaid.startDate.strftime("%m/%d/%Y");
            ws.cell(row=7, column=13).value = person.unpaid.endDate.strftime("%m/%d/%Y");

        # row 8
        ws.cell(row=8, column=1).value = "Process Level:";
        ws.cell(row=8, column=1).font = self.bold;
        ws.cell(row=8, column=2).value = int(person.processLevel);

        # Salaries
        firstCol = True;
        rowOffset = 0;
        colHead1 = 1;
        colVal1 = 2;
        colHead2 = 3;
        colVal2 = 4;
        for i in range(len(person.salaries)):
            if firstCol:
                colHead1 = 1;
                colVal1 = 2;
                colHead2 = 3;
                colVal2 = 4;
            else:
                colHead1 = 7;
                colVal1 = 9;
                colHead2 = 11;
                colVal2 = 13;

            ws.cell(row=10 + rowOffset, column=colHead1).value = "Salary " + str(i + 1);
            ws.cell(row=10 + rowOffset, column=colHead1).font = self.bold;
            ws.cell(row=10 + rowOffset, column=colHead1).border = Border(left=self.bd, top=self.bd, right=self.bd, bottom=self.bd);
            ws.cell(row=10 + rowOffset, column=colHead1).alignment = self.centerHorz;
            ws.merge_cells(start_row=10 + rowOffset, start_column=colHead1, end_row=10 + rowOffset, end_column=colVal2);

            ws.cell(row=11 + rowOffset, column=colHead1).value = "Salary:";
            ws.cell(row=11 + rowOffset, column=colHead1).font = self.bold;
            ws.cell(row=11 + rowOffset, column=colVal1).value = float(person.salaries[i].salary);
            ws.cell(row=11 + rowOffset, column=colVal1).number_format = '"$"#,##0.00_);("$"#,##0.00)';
            ws.cell(row=11 + rowOffset, column=colHead2).value = "Daily Hrs:";
            ws.cell(row=11 + rowOffset, column=colHead2).font = self.bold;
            ws.cell(row=11 + rowOffset, column=colVal2).value = person.salaries[i].dailyHours;

            ws.cell(row=12 + rowOffset, column=colHead1).value = "Contract Days:";
            ws.cell(row=12 + rowOffset, column=colHead1).font = self.bold;
            ws.cell(row=12 + rowOffset, column=colVal1).value = person.salaries[i].contractDays;
            ws.cell(row=12 + rowOffset, column=colHead2).value = "Daily Rate:";
            ws.cell(row=12 + rowOffset, column=colHead2).font = self.bold;
            ws.cell(row=12 + rowOffset, column=colVal2).value = person.salaries[i].dailyRate;
            ws.cell(row=12 + rowOffset, column=colVal2).number_format = '"$"#,##0.00_);("$"#,##0.00)';

            ws.cell(row=13 + rowOffset, column=colHead1).value = "Days at Salary:";
            ws.cell(row=13 + rowOffset, column=colHead1).font = self.bold;
            ws.cell(row=13 + rowOffset, column=colVal1).value = person.salaries[i].daysAtSalary;
            ws.cell(row=13 + rowOffset, column=colHead2).value = "Hrly Rate:";
            ws.cell(row=13 + rowOffset, column=colHead2).font = self.bold;
            ws.cell(row=13 + rowOffset, column=colVal2).value = person.salaries[i].hourlyRate;
            ws.cell(row=13 + rowOffset, column=colVal2).number_format = '"$"#,##0.00_);("$"#,##0.00)';

            if not firstCol:
                rowOffset = rowOffset + 5
            firstCol = not firstCol;

        rowOffset = rowOffset - 5;

        if firstCol:
            colHead1 = 1;
            colVal1 = 2;
            colHead2 = 3;
            colVal2 = 4;
        else:
            colHead1 = 7;
            colVal1 = 9;
            colHead2 = 11;
            colVal2 = 13;

        # row 15
        ws.cell(row=15 + rowOffset, column=colHead1).value = "Current Salary";
        ws.cell(row=15 + rowOffset, column=colHead1).border = Border(left=self.bd, top=self.bd, right=self.bd, bottom=self.bd);
        ws.cell(row=15 + rowOffset, column=colHead1).font = self.bold;
        ws.cell(row=15 + rowOffset, column=colHead1).alignment = self.centerHorz;
        ws.merge_cells(start_row=15 + rowOffset, start_column=colHead1, end_row=15 + rowOffset, end_column=colVal2);

        # row 16
        ws.cell(row=16 + rowOffset, column=colHead1).value = "Salary:";
        ws.cell(row=16 + rowOffset, column=colHead1).font = self.bold;
        ws.cell(row=16 + rowOffset, column=colVal1).value = float(person.currSalary.salary);
        ws.cell(row=16 + rowOffset, column=colVal1).number_format = '"$"#,##0.00_);("$"#,##0.00)';
        ws.cell(row=16 + rowOffset, column=colHead2).value = "Daily Hrs:";
        ws.cell(row=16 + rowOffset, column=colHead2).font = self.bold;
        ws.cell(row=16 + rowOffset, column=colVal2).value = person.currSalary.dailyHours;

        # row 17
        ws.cell(row=17 + rowOffset, column=colHead1).value = "Contract Days:";
        ws.cell(row=17 + rowOffset, column=colHead1).font = self.bold;
        ws.cell(row=17 + rowOffset, column=colVal1).value = person.currSalary.contractDays;
        ws.cell(row=17 + rowOffset, column=colHead2).value = "Daily Rate:";
        ws.cell(row=17 + rowOffset, column=colHead2).font = self.bold;
        ws.cell(row=17 + rowOffset, column=colVal2).value = person.currSalary.dailyRate;
        ws.cell(row=17 + rowOffset, column=colVal2).number_format = '"$"#,##0.00_);("$"#,##0.00)';

        # row 18
        ws.cell(row=18 + rowOffset, column=colHead1).value = "Days at Salary:";
        ws.cell(row=18 + rowOffset, column=colHead1).font = self.bold;
        ws.cell(row=18 + rowOffset, column=colVal1).value = person.currSalary.daysAtSalary;
        ws.cell(row=18 + rowOffset, column=colHead2).value = "Hrly Rate:";
        ws.cell(row=18 + rowOffset, column=colHead2).font = self.bold;
        ws.cell(row=18 + rowOffset, column=colVal2).value = person.currSalary.hourlyRate;
        ws.cell(row=18 + rowOffset, column=colVal2).number_format = '"$"#,##0.00_);("$"#,##0.00)';

        # always add the row offset
        # rowOffset = rowOffset + 5

        # row 20
        ws.cell(row=20 + rowOffset, column=1).value = "Pay Recieved";
        ws.cell(row=20 + rowOffset, column=1).alignment = self.centerHorz;
        ws.cell(row=20 + rowOffset, column=1).fill = self.grayFill;
        ws.cell(row=20 + rowOffset, column=1).font = self.whiteFont;
        ws.cell(row=20 + rowOffset, column=6).fill = self.grayFill;
        ws.cell(row=20 + rowOffset, column=7).font = self.whiteFont;
        ws.cell(row=20 + rowOffset, column=7).alignment = self.centerHorz;

        ws.merge_cells(start_row=20 + rowOffset, start_column=1, end_row=20 + rowOffset, end_column=5);
        ws.cell(row=20 + rowOffset, column=7).value = "Unearned Leave Calculation";
        ws.cell(row=20 + rowOffset, column=7).fill = self.grayFill;
        ws.merge_cells(start_row=20 + rowOffset, start_column=7, end_row=20 + rowOffset, end_column=13);

        # row 21
        ws.cell(row=21 + rowOffset, column=2).value = "Amount";
        ws.cell(row=21 + rowOffset, column=2).font = self.bold;
        ws.cell(row=21 + rowOffset, column=5).value = "Amount";
        ws.cell(row=21 + rowOffset, column=5).font = self.bold;

        # Unearned Leave Calculation
        ws.cell(row=21 + rowOffset, column=7).value = "July";
        ws.cell(row=22 + rowOffset, column=7).value = "August";
        ws.cell(row=23 + rowOffset, column=7).value = "September";
        ws.cell(row=24 + rowOffset, column=7).value = "October";
        ws.cell(row=25 + rowOffset, column=7).value = "November";
        ws.cell(row=26 + rowOffset, column=7).value = "December";
        ws.cell(row=27 + rowOffset, column=7).value = "January";
        ws.cell(row=28 + rowOffset, column=7).value = "February";
        ws.cell(row=29 + rowOffset, column=7).value = "March";
        ws.cell(row=30 + rowOffset, column=7).value = "April";
        ws.cell(row=31 + rowOffset, column=7).value = "May";
        ws.cell(row=32 + rowOffset, column=7).value = "June";
        if person.unearned.startDate != datetime.date.min:
            ws.cell(row=21 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "July");
            ws.cell(row=22 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "August");
            ws.cell(row=23 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "September");
            ws.cell(row=24 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "October");
            ws.cell(row=25 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "November");
            ws.cell(row=26 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "December");
            ws.cell(row=27 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "January");
            ws.cell(row=28 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "February");
            ws.cell(row=29 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "March");
            ws.cell(row=30 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "April");
            ws.cell(row=31 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "May");
            ws.cell(row=32 + rowOffset, column=8).value = self.numOfBusDaysInMonth(person.unearned.startDate, person.unearned.endDate, "June");

        # row 22
        ws.cell(row=22 + rowOffset, column=1).value = "Pay Dates";
        ws.cell(row=22 + rowOffset, column=1).font = self.bold;
        ws.cell(row=22 + rowOffset, column=1).border = Border(bottom=self.bd);
        ws.cell(row=22 + rowOffset, column=2).value = "Paid";
        ws.cell(row=22 + rowOffset, column=2).font = self.bold;
        ws.cell(row=22 + rowOffset, column=2).border = Border(bottom=self.bd);
        ws.cell(row=22 + rowOffset, column=4).value = "Pay Dates";
        ws.cell(row=22 + rowOffset, column=4).font = self.bold;
        ws.cell(row=22 + rowOffset, column=4).border = Border(bottom=self.bd);
        ws.cell(row=22 + rowOffset, column=5).value = "Paid";
        ws.cell(row=22 + rowOffset, column=5).font = self.bold;
        ws.cell(row=22 + rowOffset, column=5).border = Border(bottom=self.bd);

        # Pay Recieved
        salaryTimeline = []
        salaryTimeline.append(person.currSalary)
        earlyDate = datetime.date.max
        latestDate = datetime.date.min
        for i in range(len(person.salaries)):
            inserted = False;
            if person.salaries[i].date.startDate < earlyDate:
                earlyDate = person.salaries[i].date.startDate;
            if person.salaries[i].date.endDate > latestDate:
                latestDate = person.salaries[i].date.endDate
            for j in range(len(salaryTimeline)):
                if person.salaries[i].date.startDate < salaryTimeline[j].date.startDate:
                    salaryTimeline.insert(j, person.salaries[i]);
                    inserted = True;
            if not inserted:
                salaryTimeline.append(person.salaries[i]);

        totalAmount = 0;
        thisPayPeriodAmount = 0
        nextPayDate = self.firstPay;
        currDate = earlyDate if earlyDate < self.firstPay else self.firstPay;
        index = 0 if earlyDate == currDate else -1;
        weekend = [5,6];
        payPeriod = datetime.timedelta(days=14);
        row = 23 + rowOffset;
        col = 1;
        while currDate <= self.lastPay:
            # checks to see if the new salary has started
            if index + 1 < len(salaryTimeline):
                if salaryTimeline[index + 1].date.startDate == currDate:
                    index = index + 1;

            # checks to see if it is pay day
            if currDate == nextPayDate:
                if row == 36 + rowOffset:
                    row = 23 + rowOffset;
                    col = 4;
                ws.cell(row=row, column=col).value = currDate.strftime("%m/%d/%Y");
                ws.cell(row=row, column=col+1).value = thisPayPeriodAmount;
                ws.cell(row=row, column=col+1).number_format = '"$"#,##0.00_);("$"#,##0.00)';
                row = row +1;
                thisPayPeriodAmount = 0;
                nextPayDate = nextPayDate + payPeriod;

            # checks to see if it is a weekday and it's in defined salary timeline
            if currDate.weekday() not in weekend and index > -1:
                if not (currDate >= person.unpaid.startDate and currDate <= person.unpaid.endDate):
                    totalAmount = salaryTimeline[index].dailyRate + totalAmount;
                    thisPayPeriodAmount = salaryTimeline[index].dailyRate + thisPayPeriodAmount;

            # increments 1 day
            currDate = currDate + datetime.timedelta(days=1);
        print("Total Amount: " + str(totalAmount))
        ws.cell(row=37 + rowOffset, column=5).value = "=SUM(B" + str(23+rowOffset) + ":B" + str(35+rowOffset) + ")+SUM(E" + str(23+rowOffset) +":E" + str(35+rowOffset) + ")"
        ws.cell(row=37 + rowOffset, column=5).number_format = '"$"#,##0.00_);("$"#,##0.00)'

        # row 33
        ws.cell(row=33 + rowOffset, column=9).border = Border(bottom=self.bd);

        # row 34
        ws.cell(row=34 + rowOffset, column=9).value = "=SUM(I" + str(21+rowOffset) + ":I" + str(32+rowOffset) + ")";
        ws.cell(row=34 + rowOffset, column=10).value = "X";
        ws.cell(row=34 + rowOffset, column=11).value = person.currSalary.hourlyRate;
        ws.cell(row=34 + rowOffset, column=11).number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws.cell(row=34 + rowOffset, column=12).value = "=";
        ws.cell(row=34 + rowOffset, column=13).value = "=PRODUCT(K" + str(34+rowOffset) + ",I" + str(34+rowOffset) + ")";
        ws.cell(row=34 + rowOffset, column=13).number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws.cell(row=34 + rowOffset, column=9).font = self.normalBlueFont;
        ws.cell(row=34 + rowOffset, column=10).font = self.normalBlueFont;
        ws.cell(row=34 + rowOffset, column=11).font = self.normalBlueFont;
        ws.cell(row=34 + rowOffset, column=12).font = self.normalBlueFont;
        ws.cell(row=34 + rowOffset, column=13).font = self.normalBlueFont;

        # row 35
        ws.cell(row=35 + rowOffset, column=9).value = "Total";
        ws.cell(row=35 + rowOffset, column=11).value = "Cur. Hrly";
        ws.cell(row=35 + rowOffset, column=13).value = "Unearned";

        # row 36
        ws.cell(row=36 + rowOffset, column=9).value = "Unearned";
        ws.cell(row=36 + rowOffset, column=10).value = "X";
        ws.cell(row=36 + rowOffset, column=11).value = "Rate";
        ws.cell(row=36 + rowOffset, column=12).value = "=";
        ws.cell(row=36 + rowOffset, column=13).value = "Leave $";

        # row 37
        ws.cell(row=37 + rowOffset, column=1).value = "Total Amount Paid to Employee";
        ws.cell(row=37 + rowOffset, column=1).font = self.bold;
        ws.cell(row=37 + rowOffset, column=5).font = self.normalBlueFont;
#        ws.cell(row=37, column=5).value = self.totalPaid;

        # row 39
        ws.cell(row=39 + rowOffset, column=1).value = "Amount Earned";
        ws.cell(row=39 + rowOffset, column=1).font = self.bold;
        ws.cell(row=39 + rowOffset, column=7).value = "Adjustment: First Pay After the Employee Returns";
        ws.cell(row=39 + rowOffset, column=7).alignment = self.centerHorz;
        ws.cell(row=39 + rowOffset, column=7).border = Border(top=self.bd, left=self.bd, right=self.bd);
        ws.cell(row=39 + rowOffset, column=7).fill = self.grayFill;
        ws.cell(row=39 + rowOffset, column=7).font = self.whiteFont;
        ws.merge_cells(start_row=39 + rowOffset, start_column=7, end_row=39 + rowOffset, end_column=13);

        # row 40
        ws.cell(row=40 + rowOffset, column=7).value = "Earned vs";
        ws.cell(row=40 + rowOffset, column=7).alignment = self.centerHorz;
        ws.cell(row=40 + rowOffset, column=7).border = Border(left=self.bd);
        ws.cell(row=40 + rowOffset, column=9).value = "Unearned";
        ws.cell(row=40 + rowOffset, column=9).alignment = self.centerHorz;
        ws.cell(row=40 + rowOffset, column=11).value = "Gross";
        ws.cell(row=40 + rowOffset, column=11).alignment = self.centerHorz;
        ws.cell(row=40 + rowOffset, column=11).font = self.bold;
        ws.cell(row=40 + rowOffset, column=13).border = Border(right=self.bd);

        # row 41
        ws.cell(row=41 + rowOffset, column=7).value = "Received";
        ws.cell(row=41 + rowOffset, column=7).border = Border(left=self.bd);
        ws.cell(row=41 + rowOffset, column=7).alignment = self.centerHorz;
        ws.cell(row=41 + rowOffset, column=9).value = "Leave";
        ws.cell(row=41 + rowOffset, column=9).alignment = self.centerHorz;
        ws.cell(row=41 + rowOffset, column=11).value = "Adjustment";
        ws.cell(row=41 + rowOffset, column=11).font = self.bold;
        ws.cell(row=41 + rowOffset, column=11).alignment = self.centerHorz;
        ws.cell(row=41 + rowOffset, column=13).border = Border(right=self.bd);

        # row 42
        ws.cell(row=42 + rowOffset, column=7).border = Border(left=self.bd);
        ws.cell(row=42 + rowOffset, column=13).border = Border(right=self.bd);

        # row 43
        ws.cell(row=43 + rowOffset, column=7).border = Border(left=self.bd);
        ws.cell(row=43 + rowOffset, column=8).value = "+";
        ws.cell(row=43 + rowOffset, column=9).value = "=-M" + str(34 + rowOffset);
        ws.cell(row=43 + rowOffset, column=10).value = "=";
        ws.cell(row=43 + rowOffset, column=11).value = "=G" + str(43 + rowOffset) + "+I" + str(43 + rowOffset);
        ws.cell(row=43 + rowOffset, column=11).number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws.cell(row=43 + rowOffset, column=13).border = Border(right=self.bd);
        ws.cell(row=43 + rowOffset, column=7).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=8).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=9).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=10).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=11).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=12).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=13).font = self.normalBlueFont;

        rowOffsetb4 = rowOffset;
        # Amount Earned
        for i in range(len(person.salaries)):
            ws.cell(row=40 + rowOffset, column=1).value = "Salary " + str(i + 1);  # figure out this math algorithm
            ws.cell(row=40 + rowOffset, column=1).font = self.normalBlueFont;
            ws.cell(row=40 + rowOffset, column=2).value = person.salaries[i].daysAtSalary;
            ws.cell(row=40 + rowOffset, column=3).value = person.salaries[i].dailyRate;
            ws.cell(row=40 + rowOffset, column=3).number_format = '"$"#,##0.00_);("$"#,##0.00)'
            ws.cell(row=40 + rowOffset, column=4).value = "=";
            ws.cell(row=40 + rowOffset, column=5).value = float(person.salaries[i].daysAtSalary * person.salaries[i].dailyRate);
            ws.cell(row=40 + rowOffset, column=5).number_format = '"$"#,##0.00_);("$"#,##0.00)';
            ws.cell(row=40 + rowOffset, column=1).font = self.normalBlueFont;
            ws.cell(row=40 + rowOffset, column=2).font = self.normalBlueFont;
            ws.cell(row=40 + rowOffset, column=3).font = self.normalBlueFont;
            ws.cell(row=40 + rowOffset, column=4).font = self.normalBlueFont;
            ws.cell(row=40 + rowOffset, column=5).font = self.normalBlueFont;
            ws.cell(row=40 + rowOffset, column=2).alignment = self.centerHorz;
            ws.cell(row=40 + rowOffset, column=4).alignment = self.centerHorz;
            rowOffset = rowOffset + 1;
        rowOffset = rowOffset - 3 # removing teh default 3 salaries
        ws.cell(row=43 + rowOffset, column=1).value = "Current Salary";  # figure out this math algorithm
        ws.cell(row=43 + rowOffset, column=2).value = person.currSalary.daysAtSalary;
        ws.cell(row=43 + rowOffset, column=3).value = person.currSalary.dailyRate;
        ws.cell(row=43 + rowOffset, column=3).number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws.cell(row=43 + rowOffset, column=4).value = "=";
        ws.cell(row=43 + rowOffset, column=5).value = float(person.currSalary.daysAtSalary * person.currSalary.dailyRate);
        ws.cell(row=43 + rowOffset, column=5).number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws.cell(row=43 + rowOffset, column=1).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=2).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=3).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=4).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=5).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffset, column=2).alignment = self.centerHorz;
        ws.cell(row=43 + rowOffset, column=4).alignment = self.centerHorz;

        # row 44 Affected by Amount Earned
        ws.cell(row=44 + rowOffset, column=2).value = "Days Worked";
        ws.cell(row=44 + rowOffset, column=3).value = "X Daily Rate";
        totalWork = 0;
        for i in range(len(person.salaries)):
            totalWork = ws.cell(row=40 + rowOffsetb4 + i, column=5).value + totalWork;
        ws.cell(row=44 + rowOffset, column=5).value = totalWork;
        ws.cell(row=44 + rowOffset, column=5).number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws.cell(row=44 + rowOffset, column=5).border = Border(top=self.bd);
        ws.cell(row=44 + rowOffset, column=5).font = self.normalBlueFont;

        # row 46 Affected by Amount Earned
        ws.cell(row=46 + rowOffset, column=1).value = "Amount Earned Less The Amount Paid To The EE";
        ws.cell(row=46 + rowOffset, column=1).font = self.bold;
        ws.cell(row=46 + rowOffset, column=1).border = Border(top=self.bd);
        ws.cell(row=46 + rowOffset, column=2).border = Border(top=self.bd);
        ws.cell(row=46 + rowOffset, column=3).border = Border(top=self.bd);
        ws.cell(row=46 + rowOffset, column=4).border = Border(top=self.bd);
        ws.cell(row=46 + rowOffset, column=5).value = totalWork - totalAmount;
        ws.cell(row=46 + rowOffset, column=5).number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws.cell(row=46 + rowOffset, column=5).border = Border(top=self.bd, right=self.bd);
        ws.cell(row=46 + rowOffset, column=5).font = self.normalBlueFont;
        ws.cell(row=43 + rowOffsetb4, column=7).value = totalWork - totalAmount;
        ws.cell(row=46 + rowOffsetb4, column=7).number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws.cell(row=46 + rowOffsetb4, column=7).font = self.normalBlueFont

        # row 47 Affected by Amount Earned
        ws.cell(row=47 + rowOffset, column=1).value = "All employees will be adjusted to earned whether a decrease or increase.";
        ws.cell(row=47 + rowOffset, column=5).border = Border(right=self.bd);

        # row 48 Affected by Amount Earned
        ws.cell(row=48 + rowOffset, column=1).border = Border(bottom=self.bd);
        ws.cell(row=48 + rowOffset, column=2).border = Border(bottom=self.bd);
        ws.cell(row=48 + rowOffset, column=3).border = Border(bottom=self.bd);
        ws.cell(row=48 + rowOffset, column=4).border = Border(bottom=self.bd);
        ws.cell(row=48 + rowOffset, column=5).border = Border(bottom=self.bd, right=self.bd);

        # row 44
        ws.cell(row=44 + rowOffsetb4, column=7).border = Border(left=self.bd);
        ws.cell(row=44 + rowOffsetb4, column=9).value = "Translation to Hrs:";
        ws.cell(row=44 + rowOffsetb4, column=9).font = self.bold;
        ws.cell(row=44 + rowOffsetb4, column=11).value = "=K" + str(43 + rowOffsetb4) + "/M" + str(18 + rowOffsetb4);
        ws.cell(row=44 + rowOffsetb4, column=11).font = self.normalBlueFont;
        ws.cell(row=44 + rowOffsetb4, column=13).border = Border(right=self.bd);

        # row 45
        ws.cell(row=45 + rowOffsetb4,
                column=7).value = "Review Scenario  # 2 & #3 below to correctly apply this adjustment."
        ws.cell(row=45 + rowOffsetb4, column=7).border = Border(bottom=self.bd, left=self.bd, right=self.bd);
        ws.cell(row=45 + rowOffsetb4, column=7).alignment = self.centerHorz;
        ws.merge_cells(start_row=45 + rowOffsetb4, start_column=7, end_row=45 + rowOffsetb4, end_column=13);

        # row 46

        ws.cell(row=46 + rowOffsetb4, column=7).value = "Action Taken";
        ws.cell(row=46 + rowOffsetb4, column=7).font = self.bold;
        ws.cell(row=46 + rowOffsetb4, column=7).border = Border(left=self.bd);
        ws.cell(row=46 + rowOffsetb4, column=9).value = person.actionTaken;
        ws.cell(row=46 + rowOffsetb4, column=13).border = Border(right=self.bd, top=self.bd);

        # row 47

        ws.cell(row=47 + rowOffsetb4, column=7).border = Border(left=self.bd);
        ws.cell(row=47 + rowOffsetb4, column=13).border = Border(right=self.bd);
        ws.cell(row=47 + rowOffsetb4, column=13).value = self.today.strftime("%m/%d/%Y");

        # row 48
        ws.cell(row=48 + rowOffsetb4, column=7).border = Border(left=self.bd, bottom=self.bd);
        ws.cell(row=48 + rowOffsetb4, column=8).border = Border(bottom=self.bd);
        ws.cell(row=48 + rowOffsetb4, column=9).border = Border(bottom=self.bd);
        ws.cell(row=48 + rowOffsetb4, column=10).border = Border(bottom=self.bd);
        ws.cell(row=48 + rowOffsetb4, column=11).border = Border(bottom=self.bd);

        ws.cell(row=48 + rowOffsetb4, column=12).border = Border(bottom=self.bd);
        ws.cell(row=48 + rowOffsetb4, column=13).value = "Date";
        ws.cell(row=48 + rowOffsetb4, column=13).border = Border(bottom=self.bd, right=self.bd);
        ws.merge_cells(start_row=20 + rowOffsetb4, start_column=6, end_row=48 + rowOffsetb4, end_column=6);

        # save document
        wb.save(self.dest + '/' + person.name +' Payroll ' + self.today.strftime("%m.%d.%Y") + '.xlsx');
        print("saved file as: " + person.name +' Payroll ' + self.today.strftime("%m.%d.%Y") + '.xlsx');





